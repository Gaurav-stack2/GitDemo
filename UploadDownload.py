import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
driver  = webdriver.Chrome()
driver.implicitly_wait(5)

#Dynamic parameters
fruit_name = "Apple"
new_value = '870'
colName  = "price"


file_path = r"D:\Python Files\download.xlsx"

def update_excel_data(file_path, fruit_name, colName, new_value): #Update Code
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):  # for Fixed columns like price, serial no.
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i  # desired column is assigned

    for i in range(1, sheet.max_row + 1):  # this code will iterate every cell in excel and will match the desired text
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == fruit_name:
                Dict["row"] = i  # desired row is assigned


    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)






driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.find_element(By.ID,'downloadButton').click()

# edit the excel with the updated value
update_excel_data(file_path, "Apple", "price", '870')

#Upload Code
file_input = driver.find_element(By.CSS_SELECTOR,"input[type = 'file']") #any button with the upload file functionality should have the type as file.
file_input.send_keys(file_path) #send keys will choose the file and upload it automatically
#here error is because of the path syntax,so we  will use r as raw strings or double slashes

wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)") #css class name starts with. and ID with #
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
#print(f"DEBUG: actual_price = {actual_price}, new_value = {new_value}")
print(type(actual_price))
print(type(new_value))
assert  actual_price == new_value

