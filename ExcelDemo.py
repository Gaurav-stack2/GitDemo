import openpyxl
book = openpyxl.load_workbook("D:\Softwares\PythonDemo.xlsx")
sheet = book.active

Dict = {}

cell = sheet.cell(row=1, column=2) #one way of printing the value
print(cell.value)

sheet.cell(row=2,column=1).value = "Gaurav" # saving the value from python to excel

print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value) #another way of printing the value,give the cell location

for i in range(1,sheet.max_row+1):#to get all rows
    if sheet.cell(row=i,column=1).value == "TestCase2":#logic to print the data related to single test case only

        for j in range(2,sheet.max_column+1):#to get all columns
            #Dcit["lastname"] = "kumar"
           Dict[sheet.cell(row=1,column=j).value]  = sheet.cell(row=i,column=j).value


print(Dict)

book.save("D:\Softwares\PythonDemo.xlsx") #this to work we need to close the file before the script runs